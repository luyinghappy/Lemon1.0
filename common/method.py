# @Project:SCB_22
# @Auth:  禾禾
# @Time:  2021/7/20 0:02
# @E-mail:1554152079@qq.com
# @Company:测试

"""
接口自动化测试步骤：
1、excel测试用例准备OK，代码自动去读取测试数据
2、发送接口请求，得到响应结果
3、断言:执行结果vs 预期结果  --通过/不通过
4、写入最终结果到excel表格（不涉及输出测试报告）
"""
import openpyxl
import requests
from openpyxl import load_workbook

# 读取测试用例封装的函数
def read_case(filename, sheetname):
    wb = load_workbook(filename)  # 打开excel文件
    sh = wb[sheetname]  # 打开某个表单 register为表单名
    max_row = sh.max_row  # 获取总行数
    case_list = []  # 定义一个空列表
    for i in range(2, max_row + 1):  # 遍历excel
        case_dict = dict(
            case_id=sh.cell(row=i, column=1).value,  # 获取excel文件
            url=sh.cell(row=i, column=5).value,  # 获取url
            data=sh.cell(row=i, column=6).value,  # 获取data用例输入参数数据
            expect=sh.cell(row=i, column=7).value)  # 获取单元格里的expec期望
        case_list.append(case_dict)  # 每次循环，把生成dict追加到list列表中
    return case_list

# 写入结果的封装的函数
def write_result(filename, sheetname, row, column, final_result):
    wb = load_workbook(filename)  # 打开excel
    sh = wb[sheetname]  # 打开某个表单
    sh.cell(row=row, column=column).value = final_result  # 写入数
    wb.save(filename)  # 保存

# 发送接口请求封装的函数
def api_fun(url,data):
  header = {'X-Lemonban-Media-Type':'lemonban.v2','Content-Type':'application/json'}
  res = requests.post(url=url,json=data,headers=header).json() # api_fun形参url作为实参传给post()里的ulr,形参data同理
  return res  # 设置返回值

# eval() --运行被字符串包裹的Python表达式  去掉字符串的引号
"""
'data': '{"phone":"13112340000","pwd":"1234567a","rePwd":"1234567a","userName":"柠檬dls","verificationCode":"lemon"}'
"""
# str1 = '{"phone":"13112340000","pwd":"1234567a","rePwd":"1234567a","userName":"柠檬dls","verificationCode":"lemon"}'
# str2 = eval(str1)
# print(str2)









